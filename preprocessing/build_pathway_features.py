"""
build_3flag_enhanced.py
=======================
Enhanced 3-flag pathway feature builder.

Per pathway (×12) — GA-selectable:
  amp_hit         : binary — any gene in pathway amplified (CNA ≥ 2)
  del_hit         : binary — any gene in pathway deep-deleted (CNA ≤ -2)
  sv_hit          : binary — any gene in pathway with structural variant
  mut_rate_z      : z-scored — all nonsilent mutations / covered_genes
  any_rate_z      : z-scored — genes w/ (mut OR amp OR del OR SV) / covered_genes
                    (continuous, dense; rescues quartile stratification)
  zsum            : z(amp_hit) + z(del_hit) + z(sv_hit) + mut_rate_z
                    (combined continuous risk score)

Global features — GA-selectable:
  PW_HIT_COUNT       : number of pathways with ≥ 1 alteration
  PW_DRIVER_BURDEN_z : z-scored sum of (gof_rate + lof_rate) across all pathways

Fixed covariates (always included, cfg feature= line):
  TMB_log_z, TUMOR_PURITY_z, FGA_z, ANEUPLOIDY_SCORE_z
  SAMPLE_TYPE dummies (most-common = reference, dropped)
  GENE_PANEL dummies (most-common = reference, dropped)

Outputs: pathonly_3flag_enhanced/{CANCER}/{CANCER}_3flag_enh_{train,test}.csv
"""

import os, warnings
from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl
from sklearn.model_selection import train_test_split
warnings.filterwarnings('ignore')

# ── Paths ──────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / 'raw_data' / 'all_timeline_csv'
INPUT_DIR = ROOT / 'build_3flag_inputs'
CLINICAL_CSV = str(RAW_DIR / 'data_clinical_sample.csv')
CNA_CSV      = str(RAW_DIR / 'data_cna.csv')
MUT_CSV      = str(RAW_DIR / 'data_mutations.csv')
PATHWAY_XLSX = str(INPUT_DIR / 'tcga_10_pathway.xlsx')
MMC4_XLSX    = str(INPUT_DIR / '1-s2.0-S0092867418303593-mmc4.xlsx')
OUT_BASE     = str(ROOT / 'build_3flag_outputs' / 'pathonly_3flag_enhanced')

INTEG_GENOMICS = {
    'BRCA': str(INPUT_DIR / 'Breast_integrated_genomics.csv'),
    'PRAD': str(INPUT_DIR / 'Prostate_integrated_genomics.csv'),
    'LUAD': str(INPUT_DIR / 'LUAD_integrated_genomics_v3.csv'),
    'CRC':  str(INPUT_DIR / 'Colorectal_integrated_genomics.csv'),
    'PAAD': str(INPUT_DIR / 'Pancreatic_integrated_genomics.csv'),
}

# ── Mutation type sets ─────────────────────────────────────────────
TRUNCATING = {
    'Nonsense_Mutation', 'Frame_Shift_Del', 'Frame_Shift_Ins',
    'Splice_Site', 'Translation_Start_Site', 'Nonstop_Mutation',
}
MISSENSE  = {'Missense_Mutation'}
NONSILENT = TRUNCATING | MISSENSE | {'In_Frame_Del', 'In_Frame_Ins'}

# ── Gene alias normalisation ───────────────────────────────────────
ALIAS_MAP = {
    'CDKN2Ap16INK4A': 'CDKN2A', 'CDKN2Ap14ARF': 'CDKN2A',
    'MLL': 'KMT2A', 'MLL2': 'KMT2D', 'MLL3': 'KMT2C', 'MLL4': 'KMT2B',
    'FAM46C': 'TENT5C', 'WHSC1': 'NSD2', 'WHSC1L1': 'NSD3',
    'TCEB1': 'ELOC', 'C11orf30': 'EMSY',
}

# ── DDR gene set (all TSG) ─────────────────────────────────────────
DDR_TSG = {
    'BRCA1','BRCA2','PALB2','BARD1','BRIP1','RAD51B','RAD51C','RAD51D',
    'ATM','ATR','CHEK1','CHEK2','FANCA','FANCC','FANCD2','FANCE','FANCF',
    'FANCG','FANCI','FANCL','FANCM','BLM','WRN','RECQL4','NBN','MRE11',
    'RAD50','MLH1','MSH2','MSH3','MSH6','PMS2','POLE','POLD1',
    'ERCC1','ERCC2','ERCC3','ERCC4','ERCC5','XPC','XPA','MUTYH',
}

# ── Chromatin gene set ─────────────────────────────────────────────
CHROMATIN_OG = {'IDH1', 'IDH2', 'EZH2'}   # OG: missense proxy
CHROMATIN_TSG = {
    'ARID1A','ARID1B','ARID2','SMARCA4','SMARCB1','SMARCC1','SMARCC2',
    'SMARCD1','SMARCE1','ATRX','KDM5C','KDM6A',
    'KMT2A','KMT2B','KMT2C','KMT2D','SETD2','PBRM1','BAP1','NSD1','NSD2',
    'DNMT3A','TET1','TET2','EP300','CREBBP','NCOR1','NCOR2',
    'EED','SUZ12','ASXL1','ASXL2','BCOR','BCORL1',
    'MBD1','MBD2','MBD3','CHD4','CHD7','CHD8',
}
CHROMATIN_ALL = CHROMATIN_TSG | CHROMATIN_OG

# ── Pathway name mapping ───────────────────────────────────────────
PATHWAY_KEY_MAP = {
    'Cell_Cycle': 'Cell Cycle', 'HIPPO': 'HIPPO', 'MYC': 'MYC',
    'NOTCH': 'NOTCH', 'NRF2': 'NRF2', 'PI3K': 'PI3K',
    'TGF_Beta': 'TGF-Beta', 'RTK_RAS': 'RTK RAS', 'TP53': 'TP53',
    'WNT': 'WNT',
}
ALL_PW_KEYS = list(PATHWAY_KEY_MAP.keys()) + ['DDR', 'Chromatin']

# ── Cancer-specific pathway filter (filter1, from CLAUDE.md) ───────
CANCER_PATHWAYS = {
    'BRCA': ['TP53', 'PI3K', 'RTK_RAS', 'Cell_Cycle', 'Chromatin', 'DDR', 'MYC'],
    'LUAD': ['TP53', 'PI3K', 'RTK_RAS', 'Cell_Cycle', 'Chromatin', 'DDR', 'MYC', 'NOTCH', 'NRF2'],
    'PAAD': ['TP53', 'RTK_RAS', 'Cell_Cycle', 'Chromatin', 'TGF_Beta', 'MYC', 'NOTCH', 'HIPPO'],
    'PRAD': ['TP53', 'PI3K', 'Cell_Cycle', 'Chromatin', 'DDR', 'WNT'],
    'CRC':  ['TP53', 'RTK_RAS', 'WNT', 'TGF_Beta', 'PI3K', 'DDR', 'MYC', 'Chromatin'],
}


def check_mut_type(val, target_set):
    if pd.isna(val):
        return False
    return any(t.strip() in target_set for t in str(val).split(';'))


# ════════════════════════════════════════════════════════════════════
# 1. Parse reference files
# ════════════════════════════════════════════════════════════════════

def parse_mmc4(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb['Alteration level']
    header = [v for v in list(ws.iter_rows(values_only=True))[2] if v]
    return (
        {c[4:]  for c in header if c.startswith('AMP.')},
        {c[4:]  for c in header if c.startswith('DEL.')},
        {c[4:]  for c in header if c.startswith('MUT.')},
        {c[7:]  for c in header if c.startswith('FUSION.')},
    )


def parse_10pathway(xlsx_path):
    """Returns {pw_key: {'OG': set, 'TSG': set, 'all': set}}"""
    wb = openpyxl.load_workbook(xlsx_path)
    skip = {'MutSig genes', 'OncoKB-CNAs-AMP', 'OncoKB-CNAs-HOMDEL'}
    result = {}
    # pw_key = Python key (e.g. 'Cell_Cycle'), sh = Excel sheet name (e.g. 'Cell Cycle')
    for pw_key, sh in PATHWAY_KEY_MAP.items():
        if sh not in wb.sheetnames or sh in skip:
            continue
        ws = wb[sh]
        rows = list(ws.iter_rows(values_only=True))
        # Find header row (contains 'Gene' in first col)
        hdr_idx = next((i for i, r in enumerate(rows) if r[0] == 'Gene'), 1)
        og_genes, tsg_genes = set(), set()
        for row in rows[hdr_idx + 1:]:
            gene, _, role = (row[0], row[1], row[2]) if len(row) >= 3 else (None, None, None)
            if not isinstance(gene, str) or not gene.strip():
                continue
            gene = gene.strip()
            if role == 'OG':
                og_genes.add(gene)
            elif role == 'TSG':
                tsg_genes.add(gene)
        result[pw_key] = {'OG': og_genes, 'TSG': tsg_genes, 'all': og_genes | tsg_genes}
    # Add DDR and Chromatin
    result['DDR'] = {'OG': set(), 'TSG': DDR_TSG, 'all': DDR_TSG}
    result['Chromatin'] = {'OG': CHROMATIN_OG, 'TSG': CHROMATIN_TSG, 'all': CHROMATIN_ALL}
    return result


# ════════════════════════════════════════════════════════════════════
# 2. Global features: TMB, Purity, FGA, Aneuploidy
# ════════════════════════════════════════════════════════════════════

def build_global_features(sample_ids):
    """Returns DataFrame indexed by SAMPLE_ID with TMB_log, PURITY_norm, FGA, ANEUPLOIDY."""
    print("  Building global features...")

    # ── Clinical: TMB + Purity ────────────────────────────────────
    clin = pd.read_csv(CLINICAL_CSV, usecols=['SAMPLE_ID','TMB_NONSYNONYMOUS','TUMOR_PURITY'])
    clin = clin.set_index('SAMPLE_ID')
    clin['TMB_log'] = np.log1p(clin['TMB_NONSYNONYMOUS'])
    # Fill missing purity with median
    med_purity = clin['TUMOR_PURITY'].median()
    clin['TUMOR_PURITY'] = clin['TUMOR_PURITY'].fillna(med_purity)

    # ── CNA matrix: FGA + Aneuploidy ─────────────────────────────
    print("    Loading CNA matrix (gene × sample)...")
    cna = pd.read_csv(CNA_CSV, index_col=0)   # rows=genes, cols=samples

    # FGA = fraction of genes with |CNA| > 0.5
    fga = (cna.abs() > 0.5).sum(axis=0) / len(cna)
    fga.name = 'FGA'

    # CNA chromosome burden = mean(fraction of panel genes altered per chromosome)
    # Use this as a proxy for aneuploidy since IMPACT panel only reports focal CNAs
    print("    Computing chromosome CNA burden (aneuploidy proxy)...")
    mut_ref = pd.read_csv(MUT_CSV,
                          usecols=['Hugo_Symbol','Chromosome'],
                          low_memory=False).drop_duplicates('Hugo_Symbol')
    gene_chr = mut_ref.set_index('Hugo_Symbol')['Chromosome'].to_dict()
    cna_chr = pd.Series({g: gene_chr.get(g, None) for g in cna.index})
    cna_with_chr = cna.loc[cna_chr.notna()].copy()
    chr_series   = cna_chr[cna_chr.notna()]

    # Per chromosome: fraction of measured genes with any CNA (|CNA| > 0)
    # Per sample: mean of per-chromosome fractions → reflects breadth of CNA involvement
    chr_means = {}
    for chrom, genes in chr_series.groupby(chr_series).groups.items():
        sub = cna_with_chr.loc[genes]
        chr_means[chrom] = (sub.abs() > 0).mean(axis=0)   # fraction altered per sample

    chr_df = pd.DataFrame(chr_means)  # samples × chromosomes
    aneuploidy = chr_df.mean(axis=1)  # mean across chromosomes per sample
    aneuploidy.name = 'ANEUPLOIDY_SCORE'

    # ── Merge ─────────────────────────────────────────────────────
    global_df = pd.DataFrame(index=sample_ids)
    global_df = global_df.join(clin[['TMB_log', 'TUMOR_PURITY']], how='left')
    global_df = global_df.join(fga, how='left')
    global_df = global_df.join(aneuploidy, how='left')

    # Fill NaN for samples missing from CNA or clinical
    global_df['TMB_log']          = global_df['TMB_log'].fillna(global_df['TMB_log'].median())
    global_df['TUMOR_PURITY']     = global_df['TUMOR_PURITY'].fillna(med_purity)
    global_df['FGA']              = global_df['FGA'].fillna(global_df['FGA'].median())
    global_df['ANEUPLOIDY_SCORE'] = global_df['ANEUPLOIDY_SCORE'].fillna(
                                        global_df['ANEUPLOIDY_SCORE'].median())

    print(f"    Global features: {global_df.shape[1]} cols for {global_df.shape[0]} samples")
    return global_df


# ════════════════════════════════════════════════════════════════════
# 3. Per-pathway feature building
# ════════════════════════════════════════════════════════════════════

def build_pathway_features(df, pw_genes):
    """
    Vectorised pathway feature builder.
    df: long-format integrated_genomics for ONE cancer.
    Returns DataFrame indexed by SAMPLE_ID with raw (not yet z-scored) pathway features.
    """
    df = df.copy()
    df['Hugo_Symbol'] = df['Hugo_Symbol'].replace(ALIAS_MAP)

    all_sids  = df['SAMPLE_ID'].unique().tolist()
    panel_genes_map = df.groupby('GENE_PANEL')['Hugo_Symbol'].apply(set).to_dict()
    # Map each sample to its panel gene set
    sid_panel = df.drop_duplicates('SAMPLE_ID').set_index('SAMPLE_ID')['GENE_PANEL'].to_dict()

    # ── Pre-build gene-level indicator tables (sample × gene, boolean) ──
    # Pivot: for each (SAMPLE_ID, Hugo_Symbol) pair keep the WORST alteration per type
    base = df.groupby(['SAMPLE_ID', 'Hugo_Symbol']).agg(
        CNA=('CNA', 'first'),
        SV=('SV',  lambda x: x.notna().any()),
        MUT=('MUT', lambda x: ';'.join(x.dropna().astype(str)))
    ).reset_index()
    base['MUT'] = base['MUT'].replace('', np.nan)

    # ── Mutation-type booleans ──
    def any_mut_type(series, typeset):
        return series.apply(lambda v: check_mut_type(v, typeset) if pd.notna(v) else False)

    base['is_amp']        = base['CNA'].ge(2)
    base['is_del']        = base['CNA'].le(-2) & base['CNA'].notna()
    base['is_sv']         = base['SV'].astype(bool)
    base['is_nonsilent']  = any_mut_type(base['MUT'], NONSILENT)
    base['is_missense']   = any_mut_type(base['MUT'], MISSENSE)
    base['is_truncating'] = any_mut_type(base['MUT'], TRUNCATING)

    # Pivot to sample × gene for each bool column (filled with False)
    def pivot_bool(col):
        return base.pivot_table(index='SAMPLE_ID', columns='Hugo_Symbol',
                                values=col, aggfunc='max', fill_value=False).astype(bool)

    print('    Pivoting alteration tables...', flush=True)
    amp_tbl   = pivot_bool('is_amp')
    del_tbl   = pivot_bool('is_del')
    sv_tbl    = pivot_bool('is_sv')
    ns_tbl    = pivot_bool('is_nonsilent')
    miss_tbl  = pivot_bool('is_missense')
    trunc_tbl = pivot_bool('is_truncating')
    print(f'    Tables built: {amp_tbl.shape}', flush=True)

    def covered(genes, panel_genes):
        """Genes in pathway ∩ panel genes."""
        return list(genes & panel_genes)

    # ── Fully vectorised per-pathway aggregation ──────────────────
    # For each pathway: build pathway-level feature columns across ALL samples at once
    result_frames = []
    driver_burden_parts = []  # per-pathway (gof_rate + lof_rate) to sum into PW_DRIVER_BURDEN
    for pw_key in ALL_PW_KEYS:
        pdef    = pw_genes[pw_key]
        og_set  = pdef['OG']
        tsg_set = pdef['TSG']
        all_set = pdef['all']
        if pw_key == 'DDR':
            og_set, tsg_set, all_set = set(), DDR_TSG, DDR_TSG
        elif pw_key == 'Chromatin':
            og_set, tsg_set, all_set = CHROMATIN_OG, CHROMATIN_TSG, CHROMATIN_ALL

        def cols_in_tbl(tbl, genes):
            return [g for g in genes if g in tbl.columns]

        def any_hit(tbl, genes):
            c = cols_in_tbl(tbl, genes)
            if not c: return pd.Series(False, index=tbl.index)
            return tbl[c].any(axis=1)

        def count_hit(tbl, genes):
            c = cols_in_tbl(tbl, genes)
            if not c: return pd.Series(0, index=tbl.index)
            return tbl[c].sum(axis=1)

        # Per-panel denominator (all pathways same panel → compute once per panel)
        # For rate features we need per-sample denominator based on panel coverage
        # Compute per-sample denom via panel_genes_map
        sid_list = amp_tbl.index.tolist()

        def make_denom(gene_set):
            """Per-sample: number of panel genes covered for gene_set."""
            vals = [len(gene_set & panel_genes_map.get(sid_panel.get(s,''), set()))
                    for s in sid_list]
            return pd.Series(vals, index=sid_list).replace(0, 1)

        denom_all = make_denom(all_set)
        denom_og  = make_denom(og_set)
        denom_tsg = make_denom(tsg_set)

        amp_h = any_hit(amp_tbl, all_set).astype(int)
        del_h = any_hit(del_tbl, all_set).astype(int)
        sv_h  = any_hit(sv_tbl,  all_set).astype(int)
        n_ns  = count_hit(ns_tbl, all_set)
        n_gof = count_hit(miss_tbl,  og_set)
        n_lof = count_hit(trunc_tbl, tsg_set)
        driver_burden_parts.append(n_gof / denom_og + n_lof / denom_tsg)

        def any_alt_count(genes):
            c = [g for g in genes if g in ns_tbl.columns]
            if not c:
                return pd.Series(0, index=ns_tbl.index)
            cols_amp = [g for g in c if g in amp_tbl.columns]
            cols_del = [g for g in c if g in del_tbl.columns]
            cols_sv  = [g for g in c if g in sv_tbl.columns]
            union = ns_tbl[c].copy()
            if cols_amp: union[cols_amp] = union[cols_amp] | amp_tbl[cols_amp]
            if cols_del: union[cols_del] = union[cols_del] | del_tbl[cols_del]
            if cols_sv:  union[cols_sv]  = union[cols_sv]  | sv_tbl[cols_sv]
            return union.sum(axis=1)

        n_any = any_alt_count(all_set)

        pk = f'PW_{pw_key}'
        pw_frame = pd.DataFrame({
            f'{pk}_amp_hit':   amp_h,
            f'{pk}_del_hit':   del_h,
            f'{pk}_sv_hit':    sv_h,
            f'{pk}_mut_rate':  n_ns  / denom_all,
            f'{pk}_any_rate':  n_any / denom_all,
        })
        result_frames.append(pw_frame)

    pw_df = pd.concat(result_frames, axis=1).fillna(0)
    pw_df.index.name = 'SAMPLE_ID'
    # PW_DRIVER_BURDEN: sum of (gof_rate + lof_rate) across all pathways (raw, z-scored later on train)
    driver_burden = sum(driver_burden_parts)
    driver_burden.name = 'PW_DRIVER_BURDEN'
    pw_df = pw_df.join(driver_burden.to_frame())
    pw_df = pw_df.reindex(all_sids, fill_value=0)
    return pw_df


# ════════════════════════════════════════════════════════════════════
# 4. Z-score normalisation (fit on train, apply to both)
# ════════════════════════════════════════════════════════════════════

def zscore_fit_transform(train_df, test_df, rate_cols):
    """Fit z-score params on train, transform train+test. Returns (train, test, params)."""
    params = {}
    for col in rate_cols:
        mu  = train_df[col].mean()
        std = train_df[col].std()
        if std < 1e-9:
            std = 1.0
        params[col] = (mu, std)
        train_df[col + '_z'] = (train_df[col] - mu) / std
        test_df[col  + '_z'] = (test_df[col]  - mu) / std
    return train_df, test_df, params


def add_zsum(df, pw_key):
    """Add zsum = z(amp_hit) + z(del_hit) + mut_rate_z + z(sv_hit) per pathway.
    Uses pre-computed _z columns if present, otherwise raw values."""
    pk = f'PW_{pw_key}'
    parts = []
    for base in ['amp_hit', 'del_hit', 'sv_hit']:
        z_col = f'{pk}_{base}_z'
        raw_col = f'{pk}_{base}'
        if z_col in df.columns:
            parts.append(df[z_col])
        elif raw_col in df.columns:
            parts.append(df[raw_col])
    for rate_type in ['mut_rate']:
        z_col = f'{pk}_{rate_type}_z'
        if z_col in df.columns:
            parts.append(df[z_col])
    if parts:
        df[f'{pk}_zsum'] = sum(parts)
    return df


# ════════════════════════════════════════════════════════════════════
# 5. Build meta (SAMPLE_TYPE + GENE_PANEL dummies + OS)
# ════════════════════════════════════════════════════════════════════

def build_meta(df):
    meta_cols = ['SAMPLE_ID', 'SAMPLE_TYPE', 'GENE_PANEL', 'OS_status', 'OS_month']
    meta = (df[meta_cols].drop_duplicates('SAMPLE_ID')
                         .set_index('SAMPLE_ID'))
    meta = meta.rename(columns={'OS_status': 'Event_OS', 'OS_month': 'OS_MONTHS'})
    meta['Event_OS']  = meta['Event_OS'].apply(lambda x: 1 if str(x).startswith('1') else 0)
    meta['OS_MONTHS'] = meta['OS_MONTHS'].astype(float)

    # Remove unknown sample types
    meta = meta[meta['SAMPLE_TYPE'] != 'Unknown']

    for col in ['SAMPLE_TYPE', 'GENE_PANEL']:
        dummies = pd.get_dummies(meta[col], prefix=col)
        dummies.columns = [c.replace(' ', '_').replace('-', '_') for c in dummies.columns]
        # Drop most-common category as reference
        ref = dummies.sum().idxmax()
        dummies = dummies.drop(columns=[ref])
        meta = pd.concat([meta.drop(columns=[col]), dummies], axis=1)

    return meta


# ════════════════════════════════════════════════════════════════════
# 6. Main pipeline per cancer
# ════════════════════════════════════════════════════════════════════

def build_cancer(cancer, pw_genes, global_df):
    print(f"\n{'─'*60}")
    print(f"  {cancer}")

    raw_csv = INTEG_GENOMICS[cancer]
    df = pd.read_csv(raw_csv)
    df['Hugo_Symbol'] = df['Hugo_Symbol'].replace(ALIAS_MAP)
    print(f"  Rows={len(df):,}  Samples={df['SAMPLE_ID'].nunique():,}")

    # Build meta (dummies, OS)
    meta = build_meta(df)
    meta = meta[meta['OS_MONTHS'] > 0]    # remove zero-time samples

    # Build pathway features (vectorised)
    pw_raw = build_pathway_features(df, pw_genes)
    pw_raw = pw_raw.reindex(meta.index, fill_value=0)

    # Add global features
    glob = global_df.reindex(meta.index)
    pw_raw = pw_raw.join(glob, how='left')
    pw_raw = pw_raw.fillna(pw_raw.median())

    # Train/test split (stratified by Event_OS)
    sids = meta.index.tolist()
    train_sids, test_sids = train_test_split(
        sids, test_size=0.2, random_state=42, stratify=meta.loc[sids, 'Event_OS'])

    pw_train = pw_raw.loc[train_sids].copy()
    pw_test  = pw_raw.loc[test_sids].copy()

    # Z-score: rate features + global numerics (binary hits kept as-is)
    rate_cols = [c for c in pw_raw.columns
                 if any(c.endswith(s) for s in ['_mut_rate','_any_rate'])]
    global_num_cols = [c for c in ['TMB_log', 'TUMOR_PURITY', 'FGA', 'ANEUPLOIDY_SCORE',
                                   'PW_DRIVER_BURDEN']
                       if c in pw_raw.columns]
    all_zcols = rate_cols + global_num_cols

    pw_train, pw_test, zscore_params = zscore_fit_transform(
        pw_train, pw_test, [c for c in all_zcols if c in pw_train.columns])

    # Drop raw (un-z-scored) columns — both rate_cols and global_num_cols
    raw_drop = [c for c in (rate_cols + global_num_cols) if c in pw_train.columns]
    pw_train = pw_train.drop(columns=raw_drop, errors='ignore')
    pw_test  = pw_test.drop(columns=raw_drop,  errors='ignore')

    # Per-pathway zsum = z(amp_hit)+z(del_hit)+z(sv_hit)+mut_rate_z (computed after z-scoring)
    for split_df in [pw_train, pw_test]:
        for pw_key in ALL_PW_KEYS:
            add_zsum(split_df, pw_key)

    # Global multi-pathway features (computed after z-scoring)
    keep_pathways = CANCER_PATHWAYS.get(cancer, ALL_PW_KEYS)
    for split_df in [pw_train, pw_test]:
        per_pw_hit = pd.DataFrame({
            pw_key: (
                split_df.get(f'PW_{pw_key}_amp_hit', 0).astype(bool) |
                split_df.get(f'PW_{pw_key}_del_hit', 0).astype(bool) |
                split_df.get(f'PW_{pw_key}_sv_hit',  0).astype(bool) |
                (split_df.get(f'PW_{pw_key}_mut_rate_z', 0) > 0)
            ).astype(int)
            for pw_key in keep_pathways
        }, index=split_df.index)
        split_df['PW_HIT_COUNT'] = per_pw_hit.sum(axis=1)

    # Filter PW_* columns to cancer-specific pathways (filter1).
    # PW_DRIVER_BURDEN_z and PW_HIT_COUNT are global — keep as-is.
    drop_pathways = [p for p in ALL_PW_KEYS if p not in keep_pathways]
    drop_cols = []
    for split_df in [pw_train, pw_test]:
        for p in drop_pathways:
            for c in list(split_df.columns):
                if c.startswith(f'PW_{p}_'):
                    drop_cols.append(c)
        split_df.drop(columns=[c for c in set(drop_cols) if c in split_df.columns],
                      inplace=True, errors='ignore')

    # Assemble final DataFrames
    meta_end_cols   = ['Event_OS', 'OS_MONTHS']
    meta_other_cols = [c for c in meta.columns if c not in meta_end_cols]

    def assemble(pw_df, sids_list):
        meta_part = meta.loc[sids_list, meta_other_cols]
        pw_part   = pw_df.loc[sids_list]
        os_part   = meta.loc[sids_list, meta_end_cols]
        out = pd.concat([meta_part, pw_part, os_part], axis=1).reset_index(drop=True)
        return out

    train_out = assemble(pw_train, train_sids)
    test_out  = assemble(pw_test,  test_sids)

    n_pw_feats = len([c for c in train_out.columns
                      if c.startswith('PW_') or c in
                      ['TMB_log_z','TUMOR_PURITY_z','FGA_z','ANEUPLOIDY_SCORE_z',
                       'PW_HIT_COUNT','PW_DRIVER_BURDEN_z']])
    print(f"  Feature columns: {n_pw_feats} pathway/global features")
    print(f"  Train={len(train_out)}, Test={len(test_out)}")

    return train_out, test_out, zscore_params


# ════════════════════════════════════════════════════════════════════
# 7. Entry point
# ════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("build_3flag_enhanced.py")
    print("=" * 60)

    print("\n[1/4] Parsing reference files...")
    amp_genes_ref, del_genes_ref, mut_genes_ref, fusion_genes_ref = parse_mmc4(MMC4_XLSX)
    print(f"  mmc4: AMP={len(amp_genes_ref)}, DEL={len(del_genes_ref)}, "
          f"MUT={len(mut_genes_ref)}, FUSION={len(fusion_genes_ref)}")
    pw_genes = parse_10pathway(PATHWAY_XLSX)
    print(f"  Pathways: {list(pw_genes.keys())}")

    print("\n[2/4] Loading all sample IDs...")
    all_sids = set()
    for cancer, path in INTEG_GENOMICS.items():
        df_tmp = pd.read_csv(path, usecols=['SAMPLE_ID'])
        all_sids.update(df_tmp['SAMPLE_ID'].unique())
    all_sids = sorted(all_sids)
    print(f"  Total unique SAMPLE_IDs across all cancers: {len(all_sids)}")

    print("\n[3/4] Building global features (TMB, Purity, FGA, Aneuploidy)...")
    global_df = build_global_features(all_sids)

    print("\n[4/4] Building enhanced pathway features per cancer...")
    os.makedirs(OUT_BASE, exist_ok=True)

    for cancer in INTEG_GENOMICS.keys():
        try:
            train_out, test_out, zscore_params = build_cancer(cancer, pw_genes, global_df)
            out_dir = os.path.join(OUT_BASE, cancer)
            os.makedirs(out_dir, exist_ok=True)
            train_out.to_csv(f'{out_dir}/{cancer}_3flag_enh_train.csv', index=False)
            test_out.to_csv(f'{out_dir}/{cancer}_3flag_enh_test.csv',  index=False)
            zp_df = pd.DataFrame(
                [{'col': c, 'mean': mu, 'std': sd} for c, (mu, sd) in zscore_params.items()]
            )
            zp_df.to_csv(f'{out_dir}/{cancer}_zscore_params.csv', index=False)
            print(f"  ✓ Saved to {out_dir}/")
        except Exception as e:
            print(f"  ✗ {cancer} ERROR: {e}")
            import traceback; traceback.print_exc()

    print("\n=== Done ===")
    print(f"Output directory: {OUT_BASE}")

    # Print feature summary for first cancer
    import glob
    first_train = sorted(glob.glob(f'{OUT_BASE}/*/*.csv'))[0]
    df_check = pd.read_csv(first_train, nrows=1)
    pw_feats = [c for c in df_check.columns
                if c.startswith('PW_') or c in
                ['TMB_log_z','TUMOR_PURITY_z','FGA_z','ANEUPLOIDY_SCORE_z',
                 'PW_HIT_COUNT','PW_DRIVER_BURDEN_z']]
    print(f"\nFeature column summary ({len(pw_feats)} total):")
    feat_types = {
        'amp_hit':     [c for c in pw_feats if c.endswith('_amp_hit')],
        'del_hit':     [c for c in pw_feats if c.endswith('_del_hit')],
        'sv_hit':      [c for c in pw_feats if c.endswith('_sv_hit')],
        'mut_rate_z':  [c for c in pw_feats if c.endswith('_mut_rate_z')],
        'any_rate_z':  [c for c in pw_feats if c.endswith('_any_rate_z')],
        'zsum':        [c for c in pw_feats if c.endswith('_zsum')],
        'global':      [c for c in ['PW_HIT_COUNT','PW_DRIVER_BURDEN_z'] if c in pw_feats],
        'clin_fixed':  [c for c in pw_feats if c in
                        ['TMB_log_z','TUMOR_PURITY_z','FGA_z','ANEUPLOIDY_SCORE_z']],
    }
    for ftype, cols in feat_types.items():
        print(f"  {ftype:15s}: {len(cols):3d} features")


if __name__ == '__main__':
    main()
